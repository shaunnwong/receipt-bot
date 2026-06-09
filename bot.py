import base64
import json
import os
import re
from datetime import datetime, timezone

import anthropic
import openpyxl
import requests
from flask import Flask, request, jsonify

app = Flask(__name__)

BOT_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
TELEGRAM_API = f"https://api.telegram.org/bot{BOT_TOKEN}"
EXCEL_PATH = os.environ.get("EXCEL_PATH", os.path.expanduser("~/Desktop/Book1.xlsx"))
SHEET_NAME = "2026 HUAT"

claude = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])

RECEIPT_PROMPT = """Analyze this image and determine what type it is, then extract the relevant fields as JSON.

Also extract the date shown on the image itself (not today's date). Return it as "date" in DD/MM/YYYY format. If no date is visible, return null for date.

If this is a RECEIPT (paper receipt):
Return: {"type": "receipt", "date": "DD/MM/YYYY or null", "cash": <number or null>, "paynow": <number or null>, "mall_voucher": <number or null>, "qr": <number or null>}

If this is a DEVICE/SCREEN showing delivery figures:
Return: {"type": "delivery", "date": "DD/MM/YYYY or null", "grab_sales": <number or null>, "foodpanda_sales": <number or null>}

If this is a CASH DEPOSIT SLIP:
Check if account number is 3493354335 (UOB).
Return: {"type": "deposit", "date": "DD/MM/YYYY or null", "account_match": <true or false>, "account_found": "<account number on slip>", "tran_amount": <number or null>}

Return ONLY the JSON object, nothing else."""


def set_webhook(webhook_url: str):
    resp = requests.post(f"{TELEGRAM_API}/setWebhook", json={"url": webhook_url})
    resp.raise_for_status()
    return resp.json()


def download_file(file_id: str) -> bytes:
    resp = requests.get(f"{TELEGRAM_API}/getFile", params={"file_id": file_id})
    resp.raise_for_status()
    file_path = resp.json()["result"]["file_path"]
    download_url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path}"
    file_resp = requests.get(download_url)
    file_resp.raise_for_status()
    return file_resp.content


def analyze_image(image_bytes: bytes) -> dict:
    image_b64 = base64.standard_b64encode(image_bytes).decode("utf-8")
    response = claude.messages.create(
        model="claude-opus-4-8",
        max_tokens=512,
        thinking={"type": "adaptive"},
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/jpeg",
                            "data": image_b64,
                        },
                    },
                    {"type": "text", "text": RECEIPT_PROMPT},
                ],
            }
        ],
    )
    text = response.content[-1].text.strip()
    match = re.search(r'\{.*\}', text, re.DOTALL)
    return json.loads(match.group()) if match else json.loads(text)


def get_excel_row(date_str: str) -> int | None:
    try:
        photo_date = datetime.strptime(date_str, "%d/%m/%Y").date()
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb[SHEET_NAME]
        for row in range(1, 500):
            val = ws.cell(row=row, column=1).value
            if val is None:
                continue
            if isinstance(val, datetime) and val.date() == photo_date:
                wb.close()
                return row
        wb.close()
        return None
    except (ValueError, TypeError):
        return None


def update_excel(row: int, data: dict) -> str:
    wb = openpyxl.load_workbook(EXCEL_PATH, keep_links=False)
    ws = wb[SHEET_NAME]
    img_type = data.get("type")
    updates = []

    if img_type == "receipt":
        if data.get("cash") is not None:
            ws.cell(row=row, column=4).value = data["cash"]
            updates.append(f"Cash: {data['cash']}")
        if data.get("paynow") is not None:
            ws.cell(row=row, column=7).value = data["paynow"]
            updates.append(f"PayNow: {data['paynow']}")
        if data.get("mall_voucher") is not None:
            ws.cell(row=row, column=9).value = data["mall_voucher"]
            updates.append(f"Mall Voucher: {data['mall_voucher']}")
        if data.get("qr") is not None:
            ws.cell(row=row, column=11).value = data["qr"]
            updates.append(f"QR: {data['qr']}")

    elif img_type == "delivery":
        if data.get("foodpanda_sales") is not None:
            ws.cell(row=row, column=14).value = data["foodpanda_sales"]
            updates.append(f"Foodpanda Sales: {data['foodpanda_sales']}")
        if data.get("grab_sales") is not None:
            ws.cell(row=row, column=16).value = data["grab_sales"]
            updates.append(f"Grab Sales: {data['grab_sales']}")

    elif img_type == "deposit":
        if data.get("tran_amount") is not None:
            ws.cell(row=row, column=26).value = data["tran_amount"]
            updates.append(f"Bank In: {data['tran_amount']}")

    wb.save(EXCEL_PATH)
    return "\n".join(updates)


def format_reply(data: dict, excel_row: int, date_str: str, update_summary: str) -> str:
    img_type = data.get("type")
    row_info = f" (row {excel_row})" if excel_row is not None else " (could not determine row)"
    lines = [f"Date: {date_str}{row_info}"]

    if img_type == "receipt":
        lines.append("Type: Receipt")
        lines.append(f"Cash: {data.get('cash', 'N/A')}")
        lines.append(f"PayNow: {data.get('paynow', 'N/A')}")
        lines.append(f"Mall Voucher: {data.get('mall_voucher', 'N/A')}")
        lines.append(f"QR: {data.get('qr', 'N/A')}")

    elif img_type == "delivery":
        lines.append("Type: Delivery Screen")
        lines.append(f"Grab Sales: {data.get('grab_sales', 'N/A')}")
        lines.append(f"Foodpanda Sales: {data.get('foodpanda_sales', 'N/A')}")

    elif img_type == "deposit":
        lines.append("Type: Cash Deposit Slip")
        if data.get("account_match"):
            lines.append("Account: CORRECT (UOB 3493354335)")
        else:
            lines.append(f"Account: MISMATCH - found {data.get('account_found', 'unknown')}")
        lines.append(f"Tran Amount: {data.get('tran_amount', 'N/A')}")

    if excel_row is None:
        lines.append("\nCould not write to Excel: date not found on image.")
    elif update_summary:
        lines.append(f"\nUpdated Excel:\n{update_summary}")
    else:
        lines.append("\nNo values written to Excel.")

    return "\n".join(lines)


@app.route("/webhook", methods=["POST"])
def webhook():
    update = request.get_json()
    message = update.get("message", {})
    photos = message.get("photo")

    if photos:
        best = photos[-1]
        file_id = best["file_id"]
        chat_id = message["chat"]["id"]
        requests.post(f"{TELEGRAM_API}/sendMessage", json={
            "chat_id": chat_id,
            "text": "Reading your image...",
        })

        image_bytes = download_file(file_id)
        data = analyze_image(image_bytes)

        image_date = data.get("date")
        excel_row = get_excel_row(image_date) if image_date else None
        date_str = image_date or "date not found on image"

        if excel_row is not None:
            update_summary = update_excel(excel_row, data)
        else:
            update_summary = None

        reply = format_reply(data, excel_row, date_str, update_summary)

        requests.post(f"{TELEGRAM_API}/sendMessage", json={
            "chat_id": chat_id,
            "text": reply,
        })

    return jsonify({"ok": True})


if __name__ == "__main__":
    webhook_url = os.environ["WEBHOOK_URL"]
    result = set_webhook(webhook_url)
    print("Webhook set:", result)
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8080)))
